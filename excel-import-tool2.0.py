# 导入必要的库
import pandas as pd
import sys
from PyQt5.QtWidgets import *
from openpyxl import *
import openpyxl


# 创建一个窗体类，继承自QWidget
class ExcelImportTool(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        # 设置窗体的标题和大小
        self.setWindowTitle('Excel数据导入工具')
        # self.resize()

        # 定义一些变量
        self.source_file = '' # 源文件路径
        self.target_file = '' # 目标文件路径
        self.source_sheet = '' # 源文件的sheet名 
        self.target_sheet = '' # 目标文件的sheet名
        self.source_range = '' # 源文件的数据区域
        self.target_range = '' # 目标文件的数据区域
        self.source_sheet_names = ''
        self.target_sheet_names = ''

        hbox = QHBoxLayout()
        vbox = QVBoxLayout()
        # 创建一个垂直布局
        vbox1 = QVBoxLayout()
        vbox2 = QVBoxLayout()
        # 创建一个水平布局，用于放置源文件的相关控件
        hbox1 = QHBoxLayout()
        # 创建一个标签，用于显示“源文件”
        self.source_file_label = QLabel('源文件')
        self.resize(100,50)
        # 创建一个按钮，用于选择源文件
        self.source_file_button = QPushButton('选择')
        self.source_file_button.resize(100,50)
        # 绑定按钮的点击事件，调用select_source_file方法
        self.source_file_button.clicked.connect(self.select_source_file)
        # 创建一个文本框，用于显示源文件的路径
        self.source_file_lineEdit = QLineEdit()
        self.source_file_lineEdit.resize(200,50)
        # 将标签、按钮和文本框添加到水平布局中
        hbox1.addWidget(self.source_file_label)
        hbox1.addWidget(self.source_file_lineEdit)
        hbox1.addWidget(self.source_file_button)
        # 创建一个水平布局，用于放置源文件的sheet名的下拉列表
        hbox2 = QHBoxLayout()
        # 创建一个标签，用于显示“源文件的sheet名”
        self.source_sheet_label = QLabel('源文件的sheet名')
        self.source_sheet_label.resize(200,50)
        # 创建一个下拉列表，用于选择源文件的sheet名
        self.source_sheet_combox = QComboBox()
        # 绑定下拉列表的选择事件，调用select_source_sheet方法
        self.source_sheet_combox.activated[str].connect(self.select_source_sheet)
        self.source_sheet_combox.activated[str].connect(self.show_source_data)
        self.source_sheet_combox.resize(200,50)
        hbox2.addWidget(self.source_sheet_label)
        hbox2.addWidget(self.source_sheet_combox)
        # 添加选择数据区域的跳转按钮，绑定选择事件，跳转打开新的窗口
        hbox3 = QHBoxLayout()
        self.source_range_label = QLabel('请填写或选择源文件数据区域')
        self.source_range_label.resize(200,50)
        self.source_range_button = QPushButton('选择')
        self.source_range_button.clicked.connect(self.get_range)
        self.source_range_button.clicked.connect(self.source_get_excel_range)
        self.source_range_button.resize(100,50)
        # 用于显示获得的数据区域范围参数
        self.source_range_lineEdit = QLineEdit()
        self.source_range_lineEdit.resize(100,50)
        # 将标签和下拉列表添加到水平布局中
        hbox3.addWidget(self.source_range_label)
        hbox3.addWidget(self.source_range_lineEdit)
        hbox3.addWidget(self.source_range_button)
        vbox1.addLayout(hbox1)
        vbox1.addLayout(hbox2)
        vbox1.addLayout(hbox3)

        # 创建一个水平布局，用于放置目标文件的相关控件
        hbox4 = QHBoxLayout()
        # 创建一个标签，用于显示“目标文件”
        self.target_file_label = QLabel('目标文件')
        self.target_file_label.resize(100,50)
        # 创建一个按钮，用于选择目标文件
        self.target_file_button = QPushButton('选择')
        # 绑定按钮的点击事件，调用select_target_file方法
        self.target_file_button.clicked.connect(self.select_target_file)
        self.target_file_button.resize(100,50)
        # 创建一个文本框，用于显示目标文件的路径
        self.target_file_lineEdit = QLineEdit()
        self.target_file_lineEdit.resize(200,50)
        # 将标签、按钮和文本框添加到水平布局中
        hbox4.addWidget(self.target_file_label)
        hbox4.addWidget(self.target_file_lineEdit)
        hbox4.addWidget(self.target_file_button)

        # 创建一个水平布局，用于放置目标文件的sheet名的下拉列表
        hbox5 = QHBoxLayout() 
        # 创建一个标签，用于显示“目标文件的sheet名”
        self.target_sheet_label = QLabel('目标文件的sheet名')
        self.target_sheet_label.resize(200,50)
        # 创建一个下拉列表，用于选择目标文件的sheet名
        self.target_sheet_combox = QComboBox()
        self.target_sheet_combox.resize(200,50)
        # 绑定下拉列表的选择事件，调用select_target_sheet方法
        self.target_sheet_combox.activated[str].connect(self.select_target_sheet)
        self.target_sheet_combox.activated[str].connect(self.show_traget_data)
        hbox5.addWidget(self.target_sheet_label)
        hbox5.addWidget(self.target_sheet_combox)
        # 创建一个水平布局，用于放置目标文件的数据区域选择部分功能
        hbox6 = QHBoxLayout()
        # 添加选择数据区域的跳转按钮，绑定选择事件，跳转打开新的窗口
        self.target_range_lable = QLabel('请填写或选择目标文件数据区域')
        self.target_range_lable.resize(200,50)
        self.target_range_button = QPushButton('选择')
        self.target_range_button.resize(100,50)
        self.target_range_button.clicked.connect(self.get_range)
        self.target_range_button.clicked.connect(self.target_get_excel_range)
        # 用于显示获得的数据区域范围参数
        self.target_range_lineEdit = QLineEdit()
        self.target_file_lineEdit.resize(100,50)
        # 将标签和下拉列表添加到水平布局中
        hbox6.addWidget(self.target_range_lable)
        hbox6.addWidget(self.target_range_lineEdit)
        hbox6.addWidget(self.target_range_button)
        vbox2.addLayout(hbox4)
        vbox2.addLayout(hbox5)
        vbox2.addLayout(hbox6)

        hbox7 = QHBoxLayout()
        self.table = QTableWidget() # 表格控件，用于显示Excel文件的数据
        self.table.resize(800,400)
        hbox7.addWidget(self.table)
        # 创建一个水平布局

        hbox.addLayout(vbox1)
        hbox.addLayout(vbox2)
        vbox.addLayout(hbox7)
        vbox.addLayout(hbox)
        self.import_button = QPushButton('写入数据', self, clicked=self.import_data)
        self.import_button.resize(800,50)
        vbox.addWidget(self.import_button)
        # 设置窗体的布局为垂直布局
        self.setLayout(vbox)

    # 定义选择源文件的方法
    def select_source_file(self):
        # 调用QFileDialog类的getOpenFileName方法，弹出一个文件选择对话框
        # 返回值是一个元组，第一个元素是文件路径，第二个元素是文件类型
        self.source_file, _ = QFileDialog.getOpenFileName(self, '选择源文件', '.', 'Excel文件(*.xlsx)')
        # 如果文件路径不为空，就执行以下操作
        if self.source_file:
            # 将文本框的内容设置为文件路径
            self.source_file_lineEdit.setText(self.source_file)
            # 用load_workbook方法加载源文件，返回一个workbook对象
            workbook = load_workbook(self.source_file)
            # 用workbook对象的sheetnames属性，获取源文件的所有sheet名，返回一个列表
            self.source_sheet_names = workbook.sheetnames
            # 将下拉列表的值设置为源文件的sheet名列表
            self.source_sheet_combox.clear()
            self.source_sheet_combox.addItems(self.source_sheet_names)

    # 定义选择源文件的sheet名的方法
    def select_source_sheet(self, sheet):
        # 将源文件的sheet名赋值给变量
        self.source_sheet = sheet
        return self.source_sheet

    # 定义选择目标文件的方法
    def select_target_file(self):
        # 调用QFileDialog类的getOpenFileName方法，弹出一个文件选择对话框
        # 返回值是一个元组，第一个元素是文件路径，第二个元素是文件类型
        self.target_file, _ = QFileDialog.getOpenFileName(self, '选择目标文件', '.', 'Excel文件(*.xlsx)')
        # 如果文件路径不为空，就执行以下操作
        if self.target_file:
            # 将文本框的内容设置为文件路径
            self.target_file_lineEdit.setText(self.target_file)
            # 用pd.ExcelFile方法读取目标文件，返回一个ExcelFile对象
            excel_file = pd.ExcelFile(self.target_file)
            # 用ExcelFile对象的sheet_names属性，获取目标文件的所有sheet名，返回一个列表
            self.target_sheet_names = excel_file.sheet_names
            # 将下拉列表的值设置为目标文件的sheet名列表
            self.target_sheet_combox.clear()
            self.target_sheet_combox.addItems(self.target_sheet_names)

    # 定义选择目标文件的sheet名的方法
    def select_target_sheet(self, sheet):
        # 将目标文件的sheet名赋值给变量
        self.target_sheet = sheet
        return self.target_sheet
    # 在定义选择目标文件的sheet名的方法的后面，添加以下一个方法，用于读取Excel文件的数据，显示到表格控件中

    def show_source_data(self,sheet):
    # 将源文件的sheet名赋值给变量
        self.source_sheet = sheet
        # 用pd.read_excel方法读取目标文件的指定sheet的数据，返回一个DataFrame对象
        df = pd.read_excel(self.source_file, sheet_name=self.source_sheet)
        # 用DataFrame对象的shape属性，获取数据的行数和列数，赋值给变量
        rows, cols = df.shape
        # 用表格控件的setRowCount和setColumnCount方法，设置表格的行数和列数
        self.table.setRowCount(rows)
        self.table.setColumnCount(cols)
        # 用DataFrame对象的columns属性，获取数据的列名，返回一个列表
        col_names = df.columns
        # 用表格控件的setHorizontalHeaderLabels方法，设置表格的水平表头为列名
        self.table.setHorizontalHeaderLabels(col_names)
        # 用for循环，遍历数据的每一行和每一列
        for i in range(rows):
            for j in range(cols):
                # 用DataFrame对象的iloc方法，获取数据的指定位置的值，赋值给变量
                value = df.iloc[i, j]
                # 用QTableWidgetItem类，创建一个表格项，用于显示数据的值
                item = QTableWidgetItem(str(value))
                # 用表格控件的setItem方法，将表格项添加到表格的指定位置
                self.table.setItem(i, j, item)
    def show_traget_data(self,sheet):
    # 将目标文件的sheet名赋值给变量
        self.target_sheet = sheet
        # 用pd.read_excel方法读取目标文件的指定sheet的数据，返回一个DataFrame对象
        df = pd.read_excel(self.target_file, sheet_name=self.target_sheet)
        # 用DataFrame对象的shape属性，获取数据的行数和列数，赋值给变量
        rows, cols = df.shape
        # 用表格控件的setRowCount和setColumnCount方法，设置表格的行数和列数
        self.table.setRowCount(rows)
        self.table.setColumnCount(cols)
        # 用DataFrame对象的columns属性，获取数据的列名，返回一个列表
        col_names = df.columns
        # 用表格控件的setHorizontalHeaderLabels方法，设置表格的水平表头为列名
        self.table.setHorizontalHeaderLabels(col_names)
        # 用for循环，遍历数据的每一行和每一列
        for i in range(rows):
            for j in range(cols):
                # 用DataFrame对象的iloc方法，获取数据的指定位置的值，赋值给变量
                value = df.iloc[i, j]
                # 用QTableWidgetItem类，创建一个表格项，用于显示数据的值
                item = QTableWidgetItem(str(value))
                # 用表格控件的setItem方法，将表格项添加到表格的指定位置
                self.table.setItem(i, j, item)

    # 在定义读取Excel文件的数据，显示到表格控件中的方法的后面，添加以下一个方法，用于获取表格控件中选择的区域范围，并返回区域范围的参数
    def get_range(self):
        # 用表格控件的selectedRanges方法，获取表格中选择的区域范围，返回一个列表
        ranges = self.table.selectedRanges()
        # 如果列表不为空，就执行以下操作
        if ranges:
            # 用列表的第一个元素，获取选择的区域范围，赋值给变量
            range = ranges[0]
            # 用区域范围的topRow、bottomRow、leftColumn和rightColumn方法，获取选择的区域范围的上下左右的行列号，赋值给变量
            top_row = range.topRow()+1
            bottom_row = range.bottomRow()+1
            left_col = range.leftColumn()
            right_col = range.rightColumn()
            # 用区域范围的rowCount和columnCount方法，获取选择的区域范围的行数和列数，赋值给变量
            row_count = range.rowCount()
            col_count = range.columnCount()
            # 返回选择的区域范围的参数
            return ([top_row, bottom_row, left_col, right_col, row_count, col_count])
    # 将用户选择的区域范围参数转换成Excel格式的范围表示
    def source_get_excel_range(self):
        try:
        # 将元组中的值赋给变量
            top_row, bottom_row, left_col, right_col, row_count, col_count = self.get_range()
            # 使用openpyxl.utils.cell.get_column_letter函数，将列号转换成列名
            left_col_name = openpyxl.utils.cell.get_column_letter(left_col + 1)
            right_col_name = openpyxl.utils.cell.get_column_letter(right_col + 1)
            # 使用字符串的format方法，拼接列名和行号，得到区域范围的左上角和右下角的单元格的地址
            top_left_cell = '{}{}'.format(left_col_name, top_row + 1)
            bottom_right_cell = '{}{}'.format(right_col_name, bottom_row + 1)
            # 使用字符串的format方法，用冒号连接两个单元格的地址，得到区域范围的表达格式
            range_str = '{}:{}'.format(top_left_cell, bottom_right_cell)
            self.source_range_lineEdit.setText(range_str)
        except:
        # 返回区域范围的表达格式
            return range_str
    def target_get_excel_range(self):
        try:
        # 将元组中的值赋给变量
            top_row, bottom_row, left_col, right_col, row_count, col_count = self.get_range()
            # 使用openpyxl.utils.cell.get_column_letter函数，将列号转换成列名
            left_col_name = openpyxl.utils.cell.get_column_letter(left_col + 1)
            right_col_name = openpyxl.utils.cell.get_column_letter(right_col + 1)
            # 使用字符串的format方法，拼接列名和行号，得到区域范围的左上角和右下角的单元格的地址
            top_left_cell = '{}{}'.format(left_col_name, top_row + 1)
            bottom_right_cell = '{}{}'.format(right_col_name, bottom_row + 1)
            # 使用字符串的format方法，用冒号连接两个单元格的地址，得到区域范围的表达格式
            range_str = '{}:{}'.format(top_left_cell, bottom_right_cell)
            self.target_range_lineEdit.setText(range_str)
        except:
        # 返回区域范围的表达格式
            return range_str
    def import_data(self):
        wbs = load_workbook(self.source_file,data_only=True)
        wss = wbs[self.source_sheet]
        wbt = load_workbook(self.target_file,data_only=True)
        wst = wbt[self.target_sheet]
        # 定义原Excel文件的复制范围
        source_min_col,source_min_row,source_max_col,source_max_row = openpyxl.utils.cell.range_boundaries(self.source_range_lineEdit.text())

        # 定义目标Excel文件的写入其实位置
        target_start_col,target_start_row,target_max_col,target_max_row = openpyxl.utils.cell.range_boundaries(self.target_range_lineEdit.text())
        try:
            # 遍历源Excel文件的复制区域，获取每个单元格的值
            for i in range(source_min_row, source_max_row + 1):
                    for j in range(source_min_col, source_max_col + 1):
                        # 获取源Excel文件的单元格对象
                        source_cell = wss.cell(row=i, column=j)
                        # 获取源Excel文件的单元格值
                        source_value = source_cell.value
                        # 计算目标Excel文件的写入位置
                        target_row = target_start_row + i - source_min_row
                        target_col = target_start_col + j - source_min_col
                        # 获取目标Excel文件的单元格对象
                        target_cell = wst.cell(row=target_row, column=target_col)
                        # 将源Excel文件的单元格值写入目标Excel文件的单元格
                        target_cell.value = source_value
            wbt.save(self.target_file)
            QMessageBox.information(self,'提示','写入成功')
            wbs.close()
            wbt.close()
        except:
            QMessageBox.information(self,'提示','写入失败，请检查表格中是否存在合并单元格')
            wbs.close()
            wbt.close()
    
if __name__ == '__main__':
# 创建一个应用对象
    app = QApplication(sys.argv)
    # 创建一个窗体对象
    window = ExcelImportTool()
    window.setFixedSize(810,610)
    # 显示窗体
    window.show()
    # 进入应用的事件循环
    sys.exit(app.exec_())
