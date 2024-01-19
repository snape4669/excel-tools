 # 导入必要的库
import pandas as pd
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from openpyxl import *
import openpyxl


# 创建一个图形界面
window = tk.Tk()
window.title('Excel数据导入工具')
window.geometry('450x400')

# 定义一些变量和函数
source_file = '' # 源文件路径
target_file = '' # 目标文件路径
source_sheet = '' # 源文件的sheet名 
target_sheet = '' # 目标文件的sheet名
source_range = '' # 源文件的数据区域
target_range = '' # 目标文件的数据区域
source_sheet_names = ''
target_sheet_name = ''
def select_source_file():
    global source_file
    global source_sheet_name
    source_file = filedialog.askopenfilename(title='选择源文件', filetypes=[('Excel文件', '*.xlsx')])
    source_file_label.config(text=source_file)
    source_sheet_names = list(load_workbook(source_file).sheetnames)
    source_sheet_combox['values'] = source_sheet_names


def select_target_file():
    global target_file
    global target_sheet_name
    target_file = filedialog.askopenfilename(title='选择目标文件', filetypes=[('Excel文件', '*.xlsx')])
    target_file_label.config(text=target_file)
    df_t = pd.ExcelFile(target_file)
    target_sheet_name = list(df_t.sheet_names)
    target_sheet_combox['values'] = target_sheet_name

   

def import_data():
    global source_sheet
    source_sheet = source_sheet_combox.get()


    global target_sheet
    target_sheet = target_sheet_combox.get()

    global source_range
    try:
        source_range = openpyxl.utils.cell.range_boundaries(source_range_entry.get())
    except:
         success_label.config(text='请输入正确的区域范围，例如：A1:C5') 

    global target_range
    try:
        target_range = openpyxl.utils.cell.range_boundaries(target_range_entry.get())
    except:
        success_label.config(text='请输入正确的区域范围，例如：A1:C5') 

    
    wbs = load_workbook(source_file,data_only=True)
    wss = wbs[source_sheet]
    wbt = load_workbook(target_file,data_only=True)
    wst = wbt[target_sheet]
    # 定义原Excel文件的复制范围
    source_min_col,source_min_row,source_max_col,source_max_row = source_range

    # 定义目标Excel文件的写入其实位置
    target_start_col,target_start_row,target_max_col,target_max_row = target_range
    try:
        # 遍历源Excel文件的复制区域，获取每个单元格的值
        for i in range(source_min_row, source_max_row + 1):
                for j in range(source_min_col, source_max_col + 1):
                    # 获取源Excel文件的单元格对象
                    source_cell = wss.cell(row=i, column=j)
                    # 获取源Excel文件的单元格值
                    source_value = source_cell.value
                    # 计算目标Excel文件的写入位置
                    target_row = target_start_row + i - source_min_row
                    target_col = target_start_col + j - source_min_col
                    # 获取目标Excel文件的单元格对象
                    target_cell = wst.cell(row=target_row, column=target_col)
                    # 将源Excel文件的单元格值写入目标Excel文件的单元格
                    target_cell.value = source_value
        wbt.save(target_file)
        success_label.config(text='写入成功！')
        wbs.close()
        wbt.close()
    except:
        success_label.config(text='所选工作表中存在合并单元格，请取消合并后充实！')
        wbs.close()
        wbt.close()
    # except:
    # success_label.config(text="存在合并单元格，无法写入！")
    # wbs.close()
    # wbt.close()    

# 创建一些控件
left = tk.Frame()
left.grid(column=0,row=0)
right = tk.Frame()
right.grid(column=0,row=1)
bottom = tk.Frame()
bottom.grid(column=0,row=2)
source_file_label = tk.Label(left, text='')
source_file_label.grid(column=0,row=0,sticky='w',pady=10,padx=10)
source_file_button = tk.Button(left, text='请选择源文件', command=select_source_file)
source_file_button.grid(column=1,row=0,sticky='w',pady=10,padx=10)

target_file_label = tk.Label(right, text='')
target_file_label.grid(column=0,row=0,sticky='w',padx=10,pady=10)
target_file_button = tk.Button(right, text='请选择目标文件', command=select_target_file)
target_file_button.grid(column=1,row=0,sticky='w',padx=10,pady=10)

source_sheet_label = tk.Label(left, text='请选择源文件的工作表')
source_sheet_label.grid(column=0,row=1,sticky='w',padx=10,pady=10)
source_sheet_combox = ttk.Combobox(left)
source_sheet_combox.grid(column=1,row=1,sticky='w',padx=10,pady=10)


target_sheet_label = tk.Label(right, text='请选择目标文件工作表')
target_sheet_label.grid(column=0,row=1,sticky='w',padx=10,pady=10)
target_sheet_combox = ttk.Combobox(right)
target_sheet_combox.grid(column=1,row=1,sticky='w',padx=10,pady=10)


source_range_label = tk.Label(left, text='请输入源文件的数据区域（如A1:C5）')
source_range_label.grid(column=0,row=2,sticky='w',padx=10,pady=10)
source_range_entry = tk.Entry(left)
source_range_entry.grid(column=1,row=2,sticky='w',padx=10,pady=10)

target_range_label = tk.Label(right, text='请输入目标文件的数据区域（如D1:F5）')
target_range_label.grid(column=0,row=2,sticky='w',padx=10,pady=10)
target_range_entry = tk.Entry(right)
target_range_entry.grid(column=1,row=2,sticky='w',padx=10,pady=10)

import_button = tk.Button(bottom, text='导入数据', command=import_data)
import_button.grid(row=1,column=0,padx=20,pady=10)

success_label = tk.Label(bottom, text='')
success_label.grid(row=0,column=0,padx=20,pady=10)

# 运行图形界面
window.mainloop()